"""
Bulk ingest for the structured candidate + job spreadsheet format.

Supported uploads:
  - CSV, including UTF-8 BOM and latin-1 fallback
  - Excel .xlsx / .xlsm, first worksheet only

Each row may contain candidate profile columns, job requirement columns, or both.
Candidate records and job requirements are saved independently.
"""

import ast
import csv
import io
import json
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Optional

from .compliance import evaluate_compliance
from .config import RECORDS_DIR, REQUIREMENTS_DIR
from .review import add_to_queue
from .schemas import CandidateRecord, Compliance, Identity, Profile, Scores
from .store import save_record


# ---------------------------------------------------------------------------
# Safe value parsers
# ---------------------------------------------------------------------------

def _parse_list(value: Any) -> list[str]:
    """Parse a stringified Python list, or fall back to comma-split."""
    if value is None or isinstance(value, float):
        return []
    s = str(value).strip()
    if not s or s.lower() in ("none", "n/a", "nan"):
        return []
    if s.startswith("["):
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, list):
                return [
                    str(x).strip()
                    for x in parsed
                    if x and str(x).strip() and str(x).strip().lower() not in ("none", "n/a", "nan")
                ]
        except Exception:
            pass
    return [x.strip() for x in s.split(",") if x.strip() and x.strip().lower() not in ("none", "n/a", "nan")]


def _scalar(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return None if not s or s.lower() in ("none", "n/a", "nan") else s


def _first_scalar(row: dict[str, Any], *keys: str) -> Optional[str]:
    for key in keys:
        value = _scalar(row.get(key))
        if value:
            return value
    return None


def _normalise_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%b %Y")
    return str(value).strip()


def _parse_years_experience(start_dates: list[str], end_dates: list[str]) -> Optional[int]:
    """Estimate years of experience from the earliest start to latest end/now."""
    years = []
    for start in start_dates:
        try:
            for fmt in ("%b %Y", "%B %Y", "%Y"):
                try:
                    dt = datetime.strptime(start.strip(), fmt)
                    years.append(dt.year)
                    break
                except ValueError:
                    continue
        except Exception:
            continue
    if not years:
        return None
    return max(0, datetime.now().year - min(years))


def _parse_seniority(positions: list[str], experience_years: Optional[int]) -> Optional[str]:
    """Infer seniority from job title keywords or years of experience."""
    title_text = " ".join(positions).lower()
    if any(k in title_text for k in ("senior", "lead", "principal", "staff", "head", "director")):
        return "Senior"
    if any(k in title_text for k in ("junior", "fresher", "entry", "trainee", "intern")):
        return "Junior"
    if any(k in title_text for k in ("manager", "vp", "vice president")):
        return "Lead"
    if experience_years is not None:
        if experience_years >= 7:
            return "Senior"
        if experience_years >= 3:
            return "Mid"
        if experience_years >= 1:
            return "Junior"
        return "Intern"
    return None


# ---------------------------------------------------------------------------
# Column name normalization
# ---------------------------------------------------------------------------

_COLUMN_ALIASES: dict[str, str] = {
    "\ufeffjob_position_name": "job_position_name",
    "Ã¯Â»Â¿job_position_name": "job_position_name",
    "candidate": "candidate_name",
    "candidate name": "candidate_name",
    "full_name": "candidate_name",
    "full name": "candidate_name",
    "e-mail": "email",
    "email_address": "email",
    "email address": "email",
    "educationaL_requirements": "educational_requirements",
    "experiencere_requirement": "experience_requirement",
    "responsibilities.1": "job_responsibilities",
}

_CANDIDATE_COLUMNS = {
    "candidate_name",
    "name",
    "email",
    "emails",
    "address",
    "career_objective",
    "skills",
    "degree_names",
    "educational_institution_name",
    "passing_years",
    "major_field_of_studies",
    "professional_company_names",
    "positions",
    "start_dates",
    "end_dates",
    "locations",
    "languages",
    "proficiency_levels",
    "certification_skills",
    "related_skills_in_job",
}

_JOB_COLUMNS = {
    "job_position_name",
    "educational_requirements",
    "experience_requirement",
    "age_requirement",
    "job_responsibilities",
    "skills_required",
    "matched_score",
}


def _norm_header(raw: Any) -> str:
    header = _normalise_cell(raw).strip()
    alias_key = header if header in _COLUMN_ALIASES else header.lower()
    return _COLUMN_ALIASES.get(alias_key, header)


def _norm_headers(raw_headers: list[Any]) -> list[str]:
    return [_norm_header(h) for h in raw_headers]


def _rows_from_headers(headers: list[Any], values: list[list[Any]]) -> list[dict[str, str]]:
    normalised = _norm_headers(headers)
    rows = []
    for raw_row in values:
        row = {
            header: _normalise_cell(raw_row[i]) if i < len(raw_row) else ""
            for i, header in enumerate(normalised)
            if header
        }
        rows.append(row)
    return rows


def _parse_csv_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        return [], []
    rows = _rows_from_headers(headers, list(reader))
    return _norm_headers(headers), rows


def _parse_excel_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel imports require openpyxl. Install it with: pip install openpyxl") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows_iter = worksheet.iter_rows(values_only=True)
    for header_row in rows_iter:
        if any(_normalise_cell(cell) for cell in header_row):
            headers = list(header_row)
            break
    else:
        return [], []

    values = [list(row) for row in rows_iter]
    return _norm_headers(headers), _rows_from_headers(headers, values)


def _parse_upload_rows(content: bytes, filename: str) -> tuple[list[str], list[dict[str, str]], str]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        headers, rows = _parse_excel_rows(content)
        return headers, rows, "excel_import"
    if suffix == ".csv" or not suffix:
        headers, rows = _parse_csv_rows(content)
        return headers, rows, "csv_import"
    raise ValueError("Only .csv, .xlsx, and .xlsm files are accepted")


def _has_any_data(row: dict[str, Any], columns: set[str]) -> bool:
    return any(_scalar(row.get(column)) for column in columns)


# ---------------------------------------------------------------------------
# Row to CandidateRecord
# ---------------------------------------------------------------------------

def _row_to_candidate(row: dict[str, str], row_num: int) -> tuple[str, CandidateRecord]:
    """Map a single spreadsheet row to a (record_id, CandidateRecord)."""
    skills = _parse_list(row.get("skills", ""))
    degrees = _parse_list(row.get("degree_names", ""))
    institutions = _parse_list(row.get("educational_institution_name", ""))
    passing_years = _parse_list(row.get("passing_years", ""))
    majors = _parse_list(row.get("major_field_of_studies", ""))
    companies = _parse_list(row.get("professional_company_names", ""))
    positions = _parse_list(row.get("positions", ""))
    start_dates = _parse_list(row.get("start_dates", ""))
    end_dates = _parse_list(row.get("end_dates", ""))
    locations = _parse_list(row.get("locations", ""))
    languages = _parse_list(row.get("languages", ""))
    certs = _parse_list(row.get("certification_skills", ""))
    related_skills = _parse_list(row.get("related_skills_in_job", ""))
    emails = _parse_list(row.get("emails", "")) or _parse_list(row.get("email", ""))

    career_objective = _scalar(row.get("career_objective", ""))
    address = _scalar(row.get("address", ""))
    primary_name = _first_scalar(row, "candidate_name", "name") or f"Spreadsheet Candidate {row_num}"

    study_degrees: list[str] = []
    for i, deg in enumerate(degrees):
        inst = institutions[i] if i < len(institutions) else ""
        year = passing_years[i] if i < len(passing_years) else ""
        major = majors[i] if i < len(majors) else ""
        parts = [deg]
        if major and major.lower() not in ("none", "n/a"):
            parts[0] = f"{deg} ({major})"
        if inst:
            parts.append(f"at {inst}")
        if year and year.lower() not in ("n/a", "none"):
            parts.append(f"({year})")
        study_degrees.append(" ".join(parts))

    previous_jobs: list[str] = []
    for i, pos in enumerate(positions):
        co = companies[i] if i < len(companies) else ""
        previous_jobs.append(f"{pos} at {co}" if co else pos)

    yoe = _parse_years_experience(start_dates, end_dates)
    seniority = _parse_seniority(positions, yoe)
    all_skills = list(dict.fromkeys(skills + certs + related_skills))

    filled = sum([bool(primary_name), bool(career_objective), bool(all_skills), bool(previous_jobs), bool(study_degrees), bool(yoe)])
    confidence = round(min(0.40 + filled * 0.09, 0.90), 2)

    now = datetime.utcnow().isoformat() + "Z"
    record_id = f"cand_{uuid.uuid4().hex[:12]}"
    retention_until = (datetime.now(timezone.utc) + timedelta(days=730)).isoformat()

    rec = CandidateRecord(
        created_at=now,
        updated_at=now,
        identity=Identity(
            primary_name=primary_name,
            linkedin_url=None,
            emails=emails,
        ),
        profile=Profile(
            headline=positions[0] if positions else None,
            summary=career_objective,
            seniority=seniority,
            years_of_experience=yoe,
            technologies_used=all_skills,
            languages_spoken=languages,
            previous_jobs=previous_jobs,
            study_degrees=study_degrees,
            location=address or (locations[0] if locations else None),
            projects_developed=[],
        ),
        scores=Scores(extraction_confidence=confidence),
        compliance=Compliance(
            consent_basis="legitimate_interest",
            source="spreadsheet_bulk_import",
            data_region="EEA",
            retention_until=retention_until,
            human_review_required=confidence < 0.65,
        ),
    )
    return record_id, rec


# ---------------------------------------------------------------------------
# Job requirement import
# ---------------------------------------------------------------------------

def _job_fingerprint(data: dict[str, Any]) -> str:
    reqs = data.get("requirements", {})
    title = (data.get("title") or "").strip().lower()
    skills = ",".join(sorted(str(s).strip().lower() for s in reqs.get("must_have", []) if str(s).strip()))
    location = (reqs.get("location") or "").strip().lower()
    return f"{title}|{skills}|{location}"


def _existing_job_fingerprints() -> set[str]:
    fingerprints: set[str] = set()
    if not REQUIREMENTS_DIR.exists():
        return fingerprints
    for path in REQUIREMENTS_DIR.glob("*.json"):
        try:
            fingerprints.add(_job_fingerprint(json.loads(path.read_text(encoding="utf-8"))))
        except Exception:
            continue
    return fingerprints


def _row_to_job(row: dict[str, str]) -> Optional[dict[str, Any]]:
    title = _scalar(row.get("job_position_name"))
    skills_required = _parse_list(row.get("skills_required", ""))
    responsibilities = _scalar(row.get("job_responsibilities"))
    education = _scalar(row.get("educational_requirements"))
    experience = _scalar(row.get("experience_requirement"))
    age = _scalar(row.get("age_requirement"))
    if not title and not skills_required and not responsibilities:
        return None

    title = title or "Untitled spreadsheet job"
    description_parts = [
        part
        for part in (
            responsibilities,
            f"Education: {education}" if education else None,
            f"Experience: {experience}" if experience else None,
            f"Age: {age}" if age else None,
        )
        if part
    ]
    now = datetime.utcnow().isoformat() + "Z"
    req_id = f"req_{uuid.uuid4().hex[:12]}"
    return {
        "id": req_id,
        "title": title,
        "department": "Imported",
        "description": "\n".join(description_parts) or title,
        "requirements": {
            "must_have": skills_required,
            "nice_to_have": [],
            "location": None,
            "language": [],
            "category": None,
        },
        "scoring": {},
        "shortlist": [],
        "shortlist_meta": {},
        "status": "MATCHING",
        "tags": skills_required[:2],
        "created_at": now,
        "updated_at": now,
    }


def _save_job_requirement(job: dict[str, Any]) -> None:
    REQUIREMENTS_DIR.mkdir(parents=True, exist_ok=True)
    path = REQUIREMENTS_DIR / f"{job['id']}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(job, f, indent=2)


# ---------------------------------------------------------------------------
# Dedup helpers
# ---------------------------------------------------------------------------

def _candidate_fingerprint_from_parts(name: Optional[str], emails: list[str], skills: list[str], summary: Optional[str], location: Optional[str]) -> Optional[str]:
    if emails:
        return "email:" + emails[0].lower().strip()
    if name and not name.startswith("Spreadsheet Candidate "):
        return "name:" + name.lower().strip()
    if summary and len(skills) >= 3:
        return "profile:" + "|".join(sorted(skills)[:5]) + "|" + summary[:80].lower().strip() + "|" + (location or "").lower()
    return None


def _candidate_fingerprint(rec: CandidateRecord) -> Optional[str]:
    return _candidate_fingerprint_from_parts(
        rec.identity.primary_name,
        rec.identity.emails or [],
        rec.profile.technologies_used or [],
        rec.profile.summary,
        rec.profile.location,
    )


def _existing_candidate_fingerprints() -> set[str]:
    fingerprints: set[str] = set()
    if not RECORDS_DIR.exists():
        return fingerprints
    for path in RECORDS_DIR.glob("*.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            identity = data.get("identity", {})
            profile = data.get("profile", {})
            fp = _candidate_fingerprint_from_parts(
                identity.get("primary_name"),
                identity.get("emails") or [],
                profile.get("technologies_used") or [],
                profile.get("summary"),
                profile.get("location"),
            )
            if fp:
                fingerprints.add(fp)
        except Exception:
            continue
    return fingerprints


# ---------------------------------------------------------------------------
# Public ingest functions
# ---------------------------------------------------------------------------

class CSVIngestProgress:
    """Shared mutable state for background spreadsheet ingest tasks."""
    def __init__(self):
        self.total = 0
        self.rows_seen = 0
        self.processed = 0
        self.skipped = 0
        self.failed = 0
        self.jobs_created = 0
        self.errors: list[dict[str, Any]] = []
        self.done = False
        self.started_at = datetime.utcnow().isoformat() + "Z"
        self.finished_at: Optional[str] = None

    def add_error(self, row: int, error: str) -> None:
        if len(self.errors) < 20:
            self.errors.append({"row": row, "error": error[:200]})

    def to_dict(self) -> dict:
        return {
            "total": self.total,
            "rows_seen": self.rows_seen,
            "processed": self.processed,
            "skipped": self.skipped,
            "failed": self.failed,
            "jobs_created": self.jobs_created,
            "errors": self.errors[:20],
            "done": self.done,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
        }


def _finish(progress: CSVIngestProgress) -> None:
    progress.done = True
    progress.finished_at = datetime.utcnow().isoformat() + "Z"


def stream_ingest_file(content: bytes, filename: str, progress: CSVIngestProgress) -> None:
    """
    Parse and ingest a spreadsheet upload from raw bytes.

    Updates progress in-place and is designed to run in a background thread.
    """
    try:
        headers, rows, source_type = _parse_upload_rows(content, filename)
        progress.total = len(rows)

        header_set = set(headers)
        has_candidate_schema = bool(header_set & _CANDIDATE_COLUMNS)
        has_job_schema = bool(header_set & _JOB_COLUMNS)
        if not headers:
            progress.failed = 1
            progress.add_error(0, "The uploaded file is empty or has no header row.")
            return
        if not has_candidate_schema and not has_job_schema:
            progress.failed = max(len(rows), 1)
            progress.add_error(0, "No recognized candidate or job columns were found. Check the delimiter and header names.")
            return

        existing_candidates = _existing_candidate_fingerprints()
        existing_jobs = _existing_job_fingerprints()

        for row_num, row in enumerate(rows, start=1):
            progress.rows_seen += 1
            has_candidate_data = _has_any_data(row, _CANDIDATE_COLUMNS)
            has_job_data = _has_any_data(row, _JOB_COLUMNS)

            if has_job_data:
                try:
                    job = _row_to_job(row)
                    if job:
                        fp = _job_fingerprint(job)
                        if fp not in existing_jobs:
                            _save_job_requirement(job)
                            existing_jobs.add(fp)
                            progress.jobs_created += 1
                except Exception as exc:
                    progress.add_error(row_num, f"Job import failed: {exc}")

            if not has_candidate_data:
                if not has_job_data:
                    progress.skipped += 1
                continue

            try:
                record_id, rec = _row_to_candidate(row, row_num)
                fp = _candidate_fingerprint(rec)
                if fp and fp in existing_candidates:
                    progress.skipped += 1
                    continue

                event = {
                    "event_id": f"evt_{uuid.uuid4().hex[:12]}",
                    "event_type": "spreadsheet_bulk_ingest",
                    "timestamp": rec.created_at,
                    "actor": {"type": "system"},
                    "source": {"source_type": source_type, "row": row_num, "file_name": filename},
                    "changes": [{"operation": "create", "path": "/"}],
                    "confidence": rec.scores.extraction_confidence,
                }
                save_record(record_id, rec, event=event)
                if fp:
                    existing_candidates.add(fp)

                cases = evaluate_compliance(record_id)
                if cases:
                    add_to_queue(cases)

                progress.processed += 1
            except Exception as exc:
                progress.failed += 1
                progress.add_error(row_num, str(exc))
    except Exception as exc:
        progress.failed = max(progress.failed, 1)
        progress.add_error(0, str(exc))
    finally:
        _finish(progress)


def stream_ingest_csv(content: bytes, progress: CSVIngestProgress) -> None:
    """Backward-compatible wrapper for callers that only pass CSV bytes."""
    stream_ingest_file(content, "upload.csv", progress)
